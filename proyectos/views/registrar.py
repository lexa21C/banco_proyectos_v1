from django.http import JsonResponse
from proyectos.serializers.signup import UserSignUpSerializer
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from proyectos.serializers.user import UserModelSerializer  # Import the UserModelSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from proyectos.models import Perfil

@csrf_exempt
def upload_file(request):
    if request.method == 'POST':
        file = request.FILES['file']
        workbook = load_workbook(file)
        sheet = workbook.active
        perfil ={
            'documento':'',
            'tipo_documento':'',
            'rol':'',
            'usuario':'',
        }
        inscrito = {
            'perfil':'',
            'ficha':'',
        }
        users = []
        
        for row in sheet.iter_rows(values_only=True):
            row_data = {
                'username': row[0],
                'email': row[1],
                'first_name': row[2],
                'last_name': row[3],
                'password': row[4],
                'password_confirmation': row[4]
            }
            
            serializer = UserSignUpSerializer(data=row_data)
            serializer.is_valid(raise_exception=True)
            user = serializer.save()
            print(user)
            # perfil_id = Perfil.objects.create()
                        
           # Convert the User object to a dictionary using UserModelSerializer
            user_data = UserModelSerializer(user).data
            print(user_data)
            print(type(user_data))


            # per = Perfil() # instancia de un objeto perfil 
            # # print("xxxxxxxxxxxxxxxx\n",per)            
            # # per.usuario = user_data.id 

            # # como sacar los valores de los atributos de user_data SERIALIZADO JSON 
            # per.documento = user_data.docuemnto['docuemnto']       
            # per.tipo_documento = user_data.tipo_documento      
            # per.rol             = user_data.rol
            # per.usuario         = user_data.usuario
            # # x seria la instancia del objeto  per que es un objeto de Perfil  
            # x = per.save(commit = False) # crearia el id del perfil
            # x.save() # crearia el perfil asigna a la base de datos


            users.append(user_data)  # Append the serialized user data to the users list

        workbook.close()

        return JsonResponse(users, safe=False)  # Return the serialized users list as a JSON response
    
    return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
